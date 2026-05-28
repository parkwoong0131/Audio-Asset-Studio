"""Bulk authoring helpers for CSV/XLSX asset sheets."""
from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from .pipeline_helpers import read_yaml, write_yaml

ASSET_SHEET_COLUMNS = [
    "asset_id",
    "category",
    "prompt",
    "variations",
    "duration_ms",
    "format",
    "loop",
    "bpm",
    "layers",
    "intensity_layers",
    "post_process",
    "model",
    "channels",
    "sample_rate",
    "seed_farming",
    "seed_farming_keep",
    "multiplex",
    "negative_prompt",
    "cfg_scale",
    "reference_audio",
]

JSON_COLUMNS = {"layers", "intensity_layers", "post_process", "multiplex"}
INT_COLUMNS = {"variations", "duration_ms", "bpm", "sample_rate", "seed_farming", "seed_farming_keep"}
FLOAT_COLUMNS = {"cfg_scale"}
BOOL_COLUMNS = {"loop"}


def _serialize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _parse_cell(column: str, value: Any) -> Any:
    if value in ("", None):
        return None
    if column in JSON_COLUMNS:
        if isinstance(value, (list, dict)):
            return value
        return json.loads(str(value))
    if column in INT_COLUMNS:
        return int(value)
    if column in FLOAT_COLUMNS:
        return float(value)
    if column in BOOL_COLUMNS:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}
    return str(value)


def rows_from_assets(doc: dict) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for asset in doc.get("assets", []):
        row = {column: "" for column in ASSET_SHEET_COLUMNS}
        for column in ASSET_SHEET_COLUMNS:
            row[column] = _serialize_cell(asset.get(column))
        rows.append(row)
    return rows


def assets_from_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    assets: list[dict[str, Any]] = []
    for raw in rows:
        asset: dict[str, Any] = {}
        for column in ASSET_SHEET_COLUMNS:
            parsed = _parse_cell(column, raw.get(column))
            if parsed is not None:
                asset[column] = parsed
        if asset.get("asset_id") and asset.get("category"):
            assets.append(asset)
    return assets


def export_csv(doc: dict, output_path: Path) -> Path:
    rows = rows_from_assets(doc)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=ASSET_SHEET_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        return [{k: v for k, v in row.items()} for row in reader]


def export_xlsx(doc: dict, output_path: Path) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl 미설치 — XLSX export 불가") from exc

    rows = rows_from_assets(doc)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "assets"
    ws.append(ASSET_SHEET_COLUMNS)
    for row in rows:
        ws.append([row.get(column, "") for column in ASSET_SHEET_COLUMNS])
    wb.save(output_path)
    return output_path


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl 미설치 — XLSX import 불가") from exc

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header: list[str] = []
    rows: list[dict[str, Any]] = []
    for i, values in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(v or "").strip() for v in values]
            continue
        row = {}
        for key, value in zip(header, values):
            if not key:
                continue
            row[key] = value
        rows.append(row)
    return rows


def export_sheet(doc: dict, output_path: Path) -> Path:
    suffix = output_path.suffix.lower()
    if suffix == ".csv":
        return export_csv(doc, output_path)
    if suffix == ".xlsx":
        return export_xlsx(doc, output_path)
    raise ValueError(f"Unsupported sheet format: {suffix}")


def import_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix == ".xlsx":
        return read_xlsx_rows(path)
    raise ValueError(f"Unsupported import format: {suffix}")


def apply_rows_to_doc(doc: dict, rows: list[dict[str, Any]]) -> dict:
    updated = dict(doc)
    updated["assets"] = assets_from_rows(rows)
    return updated


def apply_sheet_to_yaml(sheet_path: Path, yaml_path: Path) -> dict:
    doc = read_yaml(yaml_path)
    updated = apply_rows_to_doc(doc, import_rows(sheet_path))
    write_yaml(yaml_path, updated)
    return updated


def rows_preview_text(rows: list[dict[str, Any]], limit: int = 10) -> str:
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=ASSET_SHEET_COLUMNS)
    writer.writeheader()
    writer.writerows(rows[:limit])
    return out.getvalue()
